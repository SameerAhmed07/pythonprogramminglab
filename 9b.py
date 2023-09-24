from openpyxl import Workbook
from openpyxl.styles import Font

# Create a new workbook
wb = Workbook()

# Create two sheets
sheet = wb.active
sheet.title = "Language"
wb.create_sheet(title="Capital")

# Data
lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code = ['KA', 'TS', 'TN']

# Populate "Language" sheet
sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Language"
sheet.cell(row=1, column=3).value = "Code"

ft = Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft

for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i - 2]
    sheet.cell(row=i, column=2).value = lang[i - 2]
    sheet.cell(row=i, column=3).value = code[i - 2]

# Save workbook after populating "Language" sheet
wb.save("demo.xlsx")

# Switch to "Capital" sheet
sheet = wb["Capital"]

# Populate "Capital" sheet
sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Capital"
sheet.cell(row=1, column=3).value = "Code"

ft = Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft

for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i - 2]
    sheet.cell(row=i, column=2).value = capital[i - 2]
    sheet.cell(row=i, column=3).value = code[i - 2]

# Save workbook after populating "Capital" sheet
wb.save("demo.xlsx")

# Search for capital based on state code
srchCode = input("Enter state code for finding capital: ")
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print("Corresponding capital for code", srchCode, "is", sheet.cell(row=i, column=2).value)

# Switch back to "Language" sheet
sheet = wb["Language"]

# Search for language based on state code
srchCode = input("Enter state code for finding language: ")
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print("Corresponding language for code", srchCode, "is", sheet.cell(row=i, column=2).value)

# Close the workbook
wb.close()
